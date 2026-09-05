import os
import io
import uuid
import logging
from datetime import datetime
import filetype
from fastapi import APIRouter, BackgroundTasks, UploadFile, File, HTTPException, Depends, Form, Query
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from jose import jwt, JWTError
from datetime import timedelta

from app.models import AdminDocument, PersonalDocument, BusinessDocument, User, UserRole, ChatSession, ChatMessage
from app.db import get_db, SessionLocal
import app.crud as crud
from app.auth.security import get_current_user, require_admin, SECRET_KEY, ALGORITHM
from app.utils.emailer import send_email
from app.utils.ai import extract_document_data, summarize_admin_document, extract_raw_text
from app.utils.vector_store import add_document_embedding, delete_document_embeddings, delete_user_embeddings
from app.schemas import UploadCompleteNotify, ReviewStatusUpdate, ExtractedDataUpdate
from app.limiter import limiter
from fastapi import Request

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/upload", tags=["upload"])

from pathlib import Path

# ─────────────────────────────────────────────
# Local File Storage (Contabo VPS Disk)
# ─────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parents[4]
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

def process_ai_extraction(table: str, doc_id: int, file_path: str, doc_type: str):
    import hashlib

    # Compute SHA-256 hash of the file for deduplication (Item 3)
    try:
        with open(file_path, "rb") as f:
            file_hash = hashlib.sha256(f.read()).hexdigest()
    except Exception as e:
        logger.warning(f"[AI] Could not hash file {file_path}: {e}. Proceeding without dedup.")
        file_hash = None

    db = SessionLocal()
    try:
        model_class = PersonalDocument if table == "personal" else BusinessDocument

        # Check if another doc with the same file hash already has extracted data
        if file_hash:
            cached = (
                db.query(model_class)
                .filter(
                    model_class.file_hash      == file_hash,
                    model_class.extracted_data != None,
                    model_class.id             != doc_id,
                    model_class.deleted_at     == None,
                )
                .first()
            )
            if cached:
                logger.info(
                    f"[AI] Cache hit for hash {file_hash[:12]}… — "
                    f"copying extracted_data from doc id={cached.id} (skipping API call)"
                )
                doc = db.query(model_class).filter_by(id=doc_id).first()
                if doc:
                    doc.extracted_data = cached.extracted_data
                    doc.file_hash      = file_hash
                    db.commit()
                    try:
                        raw_text = extract_raw_text(file_path)
                        if raw_text:
                            add_document_embedding(doc.id, table, doc.user_id, raw_text)
                    except Exception as ve:
                        logger.error(f"VectorStore Failed to embed {file_path}: {ve}")
                return

        # No cache — run full extraction
        extracted = extract_document_data(file_path, doc_type)
        if extracted is None:
            # Still save the hash so we don't keep retrying a file that always fails
            if file_hash:
                doc = db.query(model_class).filter_by(id=doc_id).first()
                if doc:
                    doc.file_hash = file_hash
                    db.commit()
            return

        doc = db.query(model_class).filter_by(id=doc_id).first()
        if doc:
            doc.extracted_data = extracted
            doc.file_hash      = file_hash
            db.commit()
            try:
                raw_text = extract_raw_text(file_path)
                if raw_text:
                    add_document_embedding(doc.id, table, doc.user_id, raw_text)
            except Exception as ve:
                logger.error(f"VectorStore Failed to embed {file_path}: {ve}")

    except Exception as e:
        logger.error(f"Failed to update DB with AI extraction: {e}")
    finally:
        db.close()

async def upload_to_storage(file: UploadFile) -> str:
    """Upload to local disk. Returns the filename (storage_key)."""
    await file.seek(0)
    contents = await file.read()

    # File size validation
    MAX_SIZE = 10 * 1024 * 1024  # 10 MB
    if len(contents) > MAX_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Maximum size is 10MB.")

    # Content type validation
    ALLOWED_TYPES = {
        "application/pdf",
        "image/jpeg",
        "image/png",
        "image/webp",
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }
    
    kind = filetype.guess(contents)
    sniffed_mime = kind.mime if kind else None
    
    ext = os.path.splitext(file.filename or "file")[1].lower()
    
    # Fallback for legacy .doc, .xls, and inconsistent .docx/.xlsx sniffing (which are just zip files)
    if sniffed_mime is None or sniffed_mime == "application/zip":
        if ext in [".doc", ".docx", ".xls", ".xlsx"] and file.content_type in ALLOWED_TYPES:
            sniffed_mime = file.content_type

    if sniffed_mime not in ALLOWED_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"File type not allowed or could not be verified. Allowed: PDF, JPG, PNG, WEBP, DOC, DOCX, XLS, XLSX."
        )

    ext = os.path.splitext(file.filename or "file")[1].lower()
    object_key = f"{uuid.uuid4().hex}{ext}"
    file_path = UPLOAD_DIR / object_key

    try:
        with open(file_path, "wb") as f:
            f.write(contents)
        logger.info(f"Uploaded {object_key} locally")
        return object_key
    except Exception as e:
        logger.error(f"Local upload failed: {e}")
        raise HTTPException(status_code=500, detail="File upload failed")

def delete_from_storage(storage_key: str):
    """Delete from local disk."""
    try:
        if storage_key.startswith("uploads/"):
            storage_key = storage_key.split("uploads/")[1]
        file_path = UPLOAD_DIR / storage_key
        if file_path.exists():
            file_path.unlink()
            logger.info(f"Deleted {storage_key} locally")
    except Exception as e:
        logger.warning(f"Storage delete failed for {storage_key}: {e}")

def get_presigned_url(storage_key: str) -> str:
    """Return an authenticated URL containing a short-lived JWT for viewing the file."""
    if storage_key.startswith("uploads/"):
        storage_key = storage_key.split("uploads/")[1]
        
    expire = datetime.utcnow() + timedelta(minutes=5)
    payload = {
        "sub": storage_key,
        "type": "file_access",
        "exp": expire,
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)
    return f"/api/upload/file/{storage_key}?token={token}"

@router.get("/file/{storage_key}")
def stream_file(storage_key: str, token: str = Query(...)):
    """Stream file securely using short-lived JWT token."""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        token_key = payload.get("sub")
        if token_key != storage_key or payload.get("type") != "file_access":
            raise HTTPException(status_code=403, detail="Token mismatch or invalid type")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")

    file_path = (UPLOAD_DIR / storage_key).resolve()
    
    # Path traversal protection
    if not str(file_path).startswith(str(UPLOAD_DIR.resolve()) + os.sep):
        raise HTTPException(status_code=400, detail="Invalid file key")

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(file_path)


# ─────────────────────────────────────────────
# Presigned URL — frontend "View" button
# ─────────────────────────────────────────────

@router.get("/view-url")
def get_view_url(
    key: str,
    current_user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return a presigned URL. Verifies the requesting user owns the document."""
    if not key:
        raise HTTPException(status_code=400, detail="key is required")

    # Existence check — ensure the document actually exists in the DB first
    doc = (
        db.query(PersonalDocument).filter(PersonalDocument.storage_key == key).first() or
        db.query(BusinessDocument).filter(BusinessDocument.storage_key == key).first() or
        db.query(AdminDocument).filter(AdminDocument.storage_key == key).first()
    )

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    # Ownership check — ensure this key belongs to the requesting user or user is admin
    if current_user.role not in [UserRole.admin, UserRole.super_admin] and doc.user_id != current_user.id:
        logger.warning(
            f"User {current_user.id} attempted to access storage key {key} — not owned"
        )
        raise HTTPException(status_code=403, detail="Access denied")

    url = get_presigned_url(key)
    if not url:
        raise HTTPException(status_code=500, detail="Could not generate view URL")
    return {"url": url}


# ─────────────────────────────────────────────
# Admin Documents
# ─────────────────────────────────────────────

@router.post("/admin-documents")
async def upload_admin_document(
    file: UploadFile = File(...),
    doc_key: str = Form(...),
    doc_label: str = Form(...),
    user_id: int = Form(...),
    background: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    target_user = crud.get_user_by_id(db, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")

    # Regular admins can only upload documents for user-role accounts
    if current_user.role == UserRole.admin and target_user.role != UserRole.user:
        raise HTTPException(
            status_code=403,
            detail="Admins can only upload documents for regular user accounts"
        )

    logger.info(f"Admin {current_user.id} uploading document for user {user_id}")
    storage_key = await upload_to_storage(file)

    record = AdminDocument(
        doc_key=doc_key,
        doc_label=doc_label,
        filename=file.filename,
        storage_key=storage_key,
        content_type=file.content_type,
        uploaded_by=current_user.id,
        user_id=user_id,
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    crud.log_action(db, "upload_admin_doc", user_id=current_user.id, target=f"admin_doc:{record.id}", detail=f"for user {user_id}")

    # Background task 1: AI summary generation
    def _generate_summary(doc_id: int, file_path: str, label: str):
        summary = summarize_admin_document(str(file_path), label)
        if summary:
            _db = SessionLocal()
            try:
                doc = _db.query(AdminDocument).filter_by(id=doc_id).first()
                if doc:
                    doc.ai_summary = summary
                    _db.commit()
                    logger.info(f"[Summary] Saved summary for admin_doc:{doc_id}")
            finally:
                _db.close()

    background.add_task(
        _generate_summary,
        record.id,
        str(UPLOAD_DIR / storage_key),
        doc_label,
    )

    # Background task 2: Notify client by email
    background.add_task(
        send_email,
        to=target_user.email,
        subject="New Document Uploaded — BookKeepro",
        body=f"""
        <p>Dear {target_user.name or "Sir/Ma'am"},</p>
        <p>An admin has securely uploaded a new document (<strong>{doc_label}</strong>) to your account.</p>
        <p>You can view this document by logging into your BookKeepro portal.</p>
        <br><strong>BookKeepro Team</strong>
        """,
    )

    return {
        "id":          record.id,
        "doc_label":   record.doc_label,
        "filename":    record.filename,
        "storage_key": record.storage_key,
    }


@router.get("/admin-documents")
def list_admin_documents(
    user_id: int | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if current_user.role in [UserRole.admin, UserRole.super_admin]:
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required for admin")
        docs = (
            db.query(AdminDocument)
            .filter(AdminDocument.user_id == user_id, AdminDocument.deleted_at == None)
            .order_by(AdminDocument.uploaded_at.desc())
            .all()
        )
    elif current_user.role == UserRole.user:
        docs = (
            db.query(AdminDocument)
            .filter(AdminDocument.user_id == current_user.id, AdminDocument.deleted_at == None)
            .order_by(AdminDocument.uploaded_at.desc())
            .all()
        )
    else:
        raise HTTPException(status_code=403, detail="Unauthorized")

    return [
        {
            "id":          d.id,
            "doc_key":     d.doc_key,
            "doc_label":   d.doc_label,
            "filename":    d.filename,
            "storage_key": d.storage_key,
            "created_at":  d.uploaded_at.isoformat() + "Z" if d.uploaded_at else None,
            # None = AI still processing; string = ready to show
            "ai_summary":  d.ai_summary,
        }
        for d in docs
    ]


@router.delete("/admin-documents/{doc_id}")
def delete_admin_document(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    doc = db.query(AdminDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    target_user = db.query(User).filter(User.id == doc.user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="Target user not found")

    # Scope restriction: standard admins can only delete documents for base users
    if current_user.role == UserRole.admin and target_user.role != UserRole.user:
        raise HTTPException(
            status_code=403, 
            detail="Standard admins can only manage documents for standard users"
        )

    logger.info(f"Admin {current_user.id} deleting admin document {doc_id}")
    delete_from_storage(doc.storage_key)
    doc.deleted_at = datetime.utcnow()
    db.commit()
    crud.log_action(db, "delete_admin_doc", user_id=current_user.id, target=f"admin_doc:{doc_id}")
    return {"deleted": True}


# ─────────────────────────────────────────────
# Personal Documents
# ─────────────────────────────────────────────

@router.get("/personal-documents")
def list_personal_documents(
    tax_year: int | None = None,
    skip:     int        = Query(0,   ge=0),
    limit:    int        = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(PersonalDocument).filter(
        PersonalDocument.user_id == current_user.id,
        PersonalDocument.deleted_at == None,
    )
    if tax_year:
        q = q.filter(PersonalDocument.tax_year == tax_year)
    docs = q.order_by(PersonalDocument.uploaded_at.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id":            d.id,
            "doc_type":      d.doc_type,
            "filename":      d.filename,
            "storage_key":   d.storage_key,
            "uploaded_at":   d.uploaded_at,
            "tax_year":      d.tax_year,
            "review_status": d.review_status,
            "review_note":   d.review_note,
        }
        for d in docs
    ]


@router.post("/personal-documents")
async def upload_personal_document(
    background: BackgroundTasks,
    file: UploadFile = File(...),
    doc_type: str = Form(...),
    tax_year: int = Form(2025),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not current_user.engagement_acknowledged_at:
        raise HTTPException(status_code=403, detail="You must acknowledge the engagement letter before uploading.")

    logger.info(f"User {current_user.id} uploading personal document: {file.filename}")
    storage_key = await upload_to_storage(file)

    record = PersonalDocument(
        user_id=current_user.id,
        doc_type=doc_type,
        filename=file.filename,
        storage_key=storage_key,
        content_type=file.content_type,
        tax_year=tax_year,
    )
    db.add(record)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        # Clean up the already-uploaded S3 object
        delete_from_storage(storage_key)
        raise HTTPException(
            status_code=409,
            detail="You have already uploaded this document for this tax year. "
                   "Delete the existing one to replace it."
        )
    db.refresh(record)
    crud.log_action(db, "upload_personal", user_id=current_user.id, target=f"personal_doc:{record.id}", detail=file.filename)

    admins = db.query(User).filter(User.role.in_([UserRole.admin, UserRole.super_admin])).all()
    for admin in admins:
        background.add_task(
            send_email,
            to=admin.email,
            subject="New Personal Document Uploaded — BookKeepro",
            body=f"<p><strong>{current_user.email}</strong> uploaded a personal document: <strong>{file.filename}</strong> ({doc_type}).</p>",
        )

    background.add_task(
        process_ai_extraction,
        "personal",
        record.id,
        str(UPLOAD_DIR / storage_key),
        doc_type
    )

    return {
        "id":          record.id,
        "filename":    record.filename,
        "doc_type":    record.doc_type,
        "storage_key": record.storage_key,
        "uploaded_at": record.uploaded_at,
        "tax_year":    record.tax_year,
    }


@router.delete("/personal-documents/{doc_id}")
def delete_personal_document(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    doc = (
        db.query(PersonalDocument)
        .filter(
            PersonalDocument.id == doc_id,
            PersonalDocument.user_id == current_user.id,
        )
        .first()
    )
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    logger.info(f"User {current_user.id} deleting personal document {doc_id} (soft delete)")
    # Soft delete (Option B): Keep file in storage, only mark as deleted in DB
    doc.deleted_at = datetime.utcnow()
    db.commit()
    crud.log_action(db, "delete_personal", user_id=current_user.id, target=f"personal_doc:{doc_id}")
    return {"deleted": True}


# ─────────────────────────────────────────────
# Business Documents
# ─────────────────────────────────────────────

@router.get("/business-documents")
def list_business_documents(
    tax_year: int | None = None,
    skip:     int        = Query(0,   ge=0),
    limit:    int        = Query(200, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    q = db.query(BusinessDocument).filter(
        BusinessDocument.user_id == current_user.id,
        BusinessDocument.deleted_at == None,
    )
    if tax_year:
        q = q.filter(BusinessDocument.tax_year == tax_year)
    docs = q.order_by(BusinessDocument.uploaded_at.desc()).offset(skip).limit(limit).all()
    return [
        {
            "id":            d.id,
            "doc_type":      d.business_type,
            "business_type": d.business_type,
            "filename":      d.filename,
            "storage_key":   d.storage_key,
            "uploaded_at":   d.uploaded_at,
            "tax_year":      d.tax_year,
            "review_status": d.review_status,
            "review_note":   d.review_note,
        }
        for d in docs
    ]


@router.post("/business-documents")
async def upload_business_document(
    background: BackgroundTasks,
    file: UploadFile = File(...),
    doc_type: str = Form(...),
    tax_year: int = Form(2025),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    if not current_user.engagement_acknowledged_at:
        raise HTTPException(status_code=403, detail="You must acknowledge the engagement letter before uploading.")

    logger.info(f"User {current_user.id} uploading business document: {file.filename}")
    storage_key = await upload_to_storage(file)

    record = BusinessDocument(
        user_id=current_user.id,
        business_type=doc_type,
        filename=file.filename,
        storage_key=storage_key,
        content_type=file.content_type,
        tax_year=tax_year,
    )
    db.add(record)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        delete_from_storage(storage_key)
        raise HTTPException(
            status_code=409,
            detail="You have already uploaded this document for this tax year. "
                   "Delete the existing one to replace it."
        )
    db.refresh(record)
    crud.log_action(db, "upload_business", user_id=current_user.id, target=f"business_doc:{record.id}", detail=file.filename)

    admins = db.query(User).filter(User.role.in_([UserRole.admin, UserRole.super_admin])).all()
    for admin in admins:
        background.add_task(
            send_email,
            to=admin.email,
            subject="New Business Document Uploaded — BookKeepro",
            body=f"<p><strong>{current_user.email}</strong> uploaded a business document: <strong>{file.filename}</strong> ({doc_type}).</p>",
        )
        
    background.add_task(
        process_ai_extraction,
        "business",
        record.id,
        str(UPLOAD_DIR / storage_key),
        doc_type
    )

    return {
        "id":            record.id,
        "filename":      record.filename,
        "business_type": record.business_type,
        "storage_key":   record.storage_key,
        "tax_year":      record.tax_year,
    }


@router.delete("/business-documents/{doc_id}")
def delete_business_document(
    doc_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    doc = db.query(BusinessDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if doc.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized")

    logger.info(f"User {current_user.id} deleting business document {doc_id} (soft delete)")
    # Soft delete (Option B): Keep file in storage, only mark as deleted in DB
    doc.deleted_at = datetime.utcnow()
    db.commit()
    crud.log_action(db, "delete_business", user_id=current_user.id, target=f"business_doc:{doc_id}")
    return {"deleted": True}


# ─────────────────────────────────────────────
# Review status — admin persists approve/reject decisions
# ─────────────────────────────────────────────

@router.patch("/personal-documents/{doc_id}/review-status")
def set_personal_review_status(
    doc_id: int,
    payload: ReviewStatusUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    """Admin sets approve/reject/pending on a personal document."""
    doc = db.query(PersonalDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    doc.review_status = payload.status
    doc.review_note   = payload.note
    db.commit()

    crud.log_action(
        db, f"review_{payload.status}",
        user_id=current_user.id,
        target=f"personal_doc:{doc_id}",
        detail=f"owner:{doc.user_id}" + (f" note:{payload.note}" if payload.note else ""),
    )
    return {"id": doc_id, "review_status": payload.status}


@router.patch("/business-documents/{doc_id}/review-status")
def set_business_review_status(
    doc_id: int,
    payload: ReviewStatusUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    """Admin sets approve/reject/pending on a business document."""
    doc = db.query(BusinessDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    doc.review_status = payload.status
    doc.review_note   = payload.note
    db.commit()

    crud.log_action(
        db, f"review_{payload.status}",
        user_id=current_user.id,
        target=f"business_doc:{doc_id}",
        detail=f"owner:{doc.user_id}" + (f" note:{payload.note}" if payload.note else ""),
    )
    return {"id": doc_id, "review_status": payload.status}


@router.patch("/personal-documents/{doc_id}/extracted-data")
def update_personal_extracted_data(
    doc_id: int,
    payload: ExtractedDataUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    """Admin updates the AI extracted data (JSON) for a personal document."""
    doc = db.query(PersonalDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    doc.extracted_data = payload.extracted_data
    db.commit()

    crud.log_action(db, "update_extracted_data", user_id=current_user.id, target=f"personal_doc:{doc_id}")
    return {"id": doc_id, "extracted_data": doc.extracted_data}


@router.patch("/business-documents/{doc_id}/extracted-data")
def update_business_extracted_data(
    doc_id: int,
    payload: ExtractedDataUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    """Admin updates the AI extracted data (JSON) for a business document."""
    doc = db.query(BusinessDocument).filter_by(id=doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    doc.extracted_data = payload.extracted_data
    db.commit()

    crud.log_action(db, "update_extracted_data", user_id=current_user.id, target=f"business_doc:{doc_id}")
    return {"id": doc_id, "extracted_data": doc.extracted_data}


# ─────────────────────────────────────────────
# Admin — all documents for one user
# ─────────────────────────────────────────────

@router.get("/admin/users/{user_id}/documents")
def get_user_all_documents(
    user_id: int,
    tax_year: int | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Regular admins can only view documents for user-role accounts
    if current_user.role == UserRole.admin and user.role != UserRole.user:
        raise HTTPException(
            status_code=403,
            detail="Admins can only view documents for regular user accounts"
        )

    logger.info(f"Admin {current_user.id} retrieving all documents for user {user_id}")

    pq = db.query(PersonalDocument).filter(
        PersonalDocument.user_id == user_id, PersonalDocument.deleted_at == None
    )
    bq = db.query(BusinessDocument).filter(
        BusinessDocument.user_id == user_id, BusinessDocument.deleted_at == None
    )
    if tax_year:
        pq = pq.filter(PersonalDocument.tax_year == tax_year)
        bq = bq.filter(BusinessDocument.tax_year == tax_year)

    personal_docs = pq.order_by(PersonalDocument.uploaded_at.desc()).all()
    business_docs = bq.order_by(BusinessDocument.uploaded_at.desc()).all()

    documents = [
        {
            "id":            d.id,
            "table":         "personal",
            "doc_type":      d.doc_type,
            "filename":      d.filename,
            "storage_key":   d.storage_key,
            "uploaded_at":   d.uploaded_at.isoformat() + "Z" if d.uploaded_at else None,
            "tax_year":      d.tax_year,
            "review_status": d.review_status,
            "review_note":   d.review_note,
            "extracted_data": d.extracted_data,
        }
        for d in personal_docs
    ] + [
        {
            "id":            d.id,
            "table":         "business",
            "doc_type":      d.business_type,
            "filename":      d.filename,
            "storage_key":   d.storage_key,
            "uploaded_at":   d.uploaded_at.isoformat() + "Z" if d.uploaded_at else None,
            "tax_year":      d.tax_year,
            "review_status": d.review_status,
            "review_note":   d.review_note,
            "extracted_data": d.extracted_data,
        }
        for d in business_docs
    ]

    return {
        "user": {
            "id":    user.id,
            "name":  user.name,
            "email": user.email,
            "role":  user.role.value if hasattr(user.role, "value") else str(user.role),
            "engagement_acknowledged_at": (
                user.engagement_acknowledged_at.isoformat()
                if user.engagement_acknowledged_at else None
            ),
        },
        "documents": documents,
    }


# ─────────────────────────────────────────────
# Admin — delete a user and all documents
# ─────────────────────────────────────────────

@router.delete("/admin/users/{user_id}")
def delete_user_completely(
    user_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Regular admins can only delete user-role accounts
    # Super_admin can delete admin or user accounts (but not other super_admins)
    if current_user.role == UserRole.admin and user.role != UserRole.user:
        raise HTTPException(
            status_code=403,
            detail="Admins can only delete regular user accounts"
        )
    if current_user.role == UserRole.super_admin and user.role == UserRole.super_admin:
        raise HTTPException(
            status_code=403,
            detail="Super admin accounts cannot be deleted through the dashboard"
        )

    logger.info(f"Admin {current_user.id} deleting user {user_id}")

    try:
        for doc in db.query(AdminDocument).filter(AdminDocument.user_id == user_id).all():
            delete_from_storage(doc.storage_key)
            db.delete(doc)

        for doc in db.query(AdminDocument).filter(AdminDocument.uploaded_by == user_id).all():
            delete_from_storage(doc.storage_key)
            db.delete(doc)

        for doc in db.query(PersonalDocument).filter(PersonalDocument.user_id == user_id).all():
            delete_from_storage(doc.storage_key)
            db.delete(doc)

        for doc in db.query(BusinessDocument).filter(BusinessDocument.user_id == user_id).all():
            delete_from_storage(doc.storage_key)
            db.delete(doc)

        # Clean up legacy uploaded_files rows to avoid FK constraint errors on db.delete(user)
        from app.models import UploadedFile
        db.query(UploadedFile).filter(UploadedFile.owner_id == user_id).delete()

        # Log before deleting the user (audit_logs.user_id FK is SET NULL, so logs survive)
        crud.log_action(db, "delete_user", user_id=current_user.id, target=f"user:{user_id}", detail=f"{user.email}")

        db.delete(user)
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"User deletion failed for {user_id}: {e}")
        raise HTTPException(status_code=500, detail="Deletion failed. No changes were made.")

    return {"deleted": True, "user_id": user_id}


# ─────────────────────────────────────────────
# Batch upload notification — call once after all files uploaded (#11)
# ─────────────────────────────────────────────

@router.post("/notify-upload-complete")
@limiter.limit("5/minute")
async def notify_upload_complete(
    request: Request,
    payload: UploadCompleteNotify,
    background: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Send one summary email after a user finishes uploading all documents.
    Call this once from the frontend after the upload session is complete.
    Replaces the old per-file email approach.
    """
    admin_email = os.getenv("ADMIN_EMAIL")
    count       = payload.file_count
    category    = payload.doc_category

    background.add_task(
        send_email,
        to=current_user.email,
        subject=f"{category} Documents Uploaded — BookKeepro",
        body=f"""
        <p>Dear {current_user.name or "Sir/Ma'am"},</p>
        <p>We received <strong>{count}</strong> {category.lower()}
           document{"s" if count != 1 else ""} from you.</p>
        <p>Our team will review {"them" if count != 1 else "it"} and update you shortly.</p>
        <br><strong>BookKeepro Team</strong>
        """,
    )
    admins = db.query(User).filter(User.role.in_([UserRole.admin, UserRole.super_admin])).all()
    for admin in admins:
        background.add_task(
            send_email,
            to=admin.email,
            subject=f"New {category} Upload — BookKeepro",
            body=f"<p><strong>{current_user.email}</strong> uploaded "
                 f"{count} {category.lower()} document{'s' if count != 1 else ''}.</p>",
        )
    return {"status": "notified"}


from pydantic import BaseModel

class NotifyApprovalRequest(BaseModel):
    user_id: int

@router.post("/notify/personal")
async def notify_personal_approval(
    payload: NotifyApprovalRequest,
    background: BackgroundTasks,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    user = db.query(User).filter(User.id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    background.add_task(
        send_email,
        to=user.email,
        subject="Personal Documents Approved — BookKeepro",
        body=f"""
        <p>Dear {user.name or "Sir/Ma'am"},</p>
        <p>Your personal documents have been successfully reviewed and approved.</p>
        <br><strong>BookKeepro Team</strong>
        """,
    )
    return {"status": "email_sent"}

@router.post("/notify/business")
async def notify_business_approval(
    payload: NotifyApprovalRequest,
    background: BackgroundTasks,
    db: Session = Depends(get_db),
    _=Depends(require_admin),
):
    user = db.query(User).filter(User.id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    background.add_task(
        send_email,
        to=user.email,
        subject="Business Documents Approved — BookKeepro",
        body=f"""
        <p>Dear {user.name or "Sir/Ma'am"},</p>
        <p>Your business documents have been successfully reviewed and approved.</p>
        <br><strong>BookKeepro Team</strong>
        """,
    )
    return {"status": "email_sent"}

# ─────────────────────────────────────────────
# Required Document Templates
# ─────────────────────────────────────────────

from app.models import RequiredDocumentTemplate

@router.get("/templates")
def list_templates(
    category: str,
    tax_year: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    templates = (
        db.query(RequiredDocumentTemplate)
        .filter_by(category=category, tax_year=tax_year)
        .order_by(RequiredDocumentTemplate.id.asc())
        .all()
    )

    results = []
    for t in templates:
        # Resolve the raw storage key, self-healing old rows that stored a
        # pre-baked presigned URL in file_url instead of a bare filename.
        raw_key = t.storage_key
        if not raw_key and t.file_url:
            # Legacy row: extract the filename from a URL like
            # "/api/upload/file/abc123.pdf?token=..."
            try:
                raw_key = t.file_url.split("/api/upload/file/")[1].split("?")[0]
                # Persist the self-healed key so next read skips this path
                t.storage_key = raw_key
                db.commit()
            except (IndexError, AttributeError):
                raw_key = None

        results.append({
            "id": t.id,
            "name": t.name,
            # Fresh presigned URL generated on every read — never stale
            # Fallback to legacy file_url for static paths like /images/...
            "download": get_presigned_url(raw_key) if raw_key else t.file_url,
        })
    return results

@router.post("/admin/templates")
async def upload_template(
    file: UploadFile = File(None),
    category: str = Form(...),
    tax_year: int = Form(...),
    name: str = Form(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    # Only super admin can manage templates
    if current_user.role != UserRole.super_admin:
        raise HTTPException(status_code=403, detail="Only super admins can manage templates")

    raw_key = None
    if file and file.filename:
        raw_key = await upload_to_storage(file)  # raw filename, e.g. "abc123.pdf"

    record = RequiredDocumentTemplate(
        category=category,
        tax_year=tax_year,
        name=name,
        storage_key=raw_key,   # store raw key; fresh URLs generated on every read
        file_url=None,         # no longer used for new rows
    )

    try:
        db.add(record)
        db.commit()
        db.refresh(record)
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=400, detail="A template with this name already exists for the given year.")

    return {
        "id": record.id,
        "name": record.name,
        "download": get_presigned_url(record.storage_key) if record.storage_key else None,
    }

@router.delete("/admin/templates/{template_id}")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    if current_user.role != UserRole.super_admin:
        raise HTTPException(status_code=403, detail="Only super admins can manage templates")

    record = db.query(RequiredDocumentTemplate).filter_by(id=template_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Template not found")

    # Delete the physical file — prefer storage_key, fall back to legacy file_url path
    key_to_delete = record.storage_key
    if not key_to_delete and record.file_url:
        try:
            key_to_delete = record.file_url.split("/api/upload/file/")[1].split("?")[0]
        except (IndexError, AttributeError):
            key_to_delete = None

    if key_to_delete:
        delete_from_storage(key_to_delete)

    db.delete(record)
    db.commit()
    return {"deleted": True}

from fastapi.responses import StreamingResponse
import csv
import io
import json

@router.get("/admin/users/{user_id}/export")
def export_user_documents_excel(
    user_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
    _=Depends(require_admin),
):
    """Export all documents for a user as a formatted Excel (.xlsx) workbook."""
    import io
    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if current_user.role == UserRole.admin and user.role != UserRole.user:
        raise HTTPException(status_code=403, detail="Admins can only export documents for regular user accounts")

    personal_docs = db.query(PersonalDocument).filter(
        PersonalDocument.user_id == user_id, PersonalDocument.deleted_at == None
    ).all()
    business_docs = db.query(BusinessDocument).filter(
        BusinessDocument.user_id == user_id, BusinessDocument.deleted_at == None
    ).all()

    # Collect all extracted_data keys to build dynamic columns
    extracted_keys = set()
    all_docs = []

    for d in personal_docs:
        row = {
            "ID": f"P-{d.id}",
            "Table": "Personal",
            "Doc Type": d.doc_type,
            "Filename": d.filename,
            "Uploaded At": d.uploaded_at.strftime("%Y-%m-%d %H:%M") if d.uploaded_at else "",
            "Tax Year": d.tax_year or "",
            "Review Status": (d.review_status or "").capitalize(),
            "Review Note": d.review_note or "",
        }
        if d.extracted_data:
            try:
                data = _json.loads(d.extracted_data) if isinstance(d.extracted_data, str) else d.extracted_data
                if isinstance(data, dict):
                    for k, v in data.items():
                        col = f"AI: {k}"
                        extracted_keys.add(col)
                        row[col] = str(v)
            except Exception:
                pass
        all_docs.append(row)

    for d in business_docs:
        row = {
            "ID": f"B-{d.id}",
            "Table": "Business",
            "Doc Type": d.business_type,
            "Filename": d.filename,
            "Uploaded At": d.uploaded_at.strftime("%Y-%m-%d %H:%M") if d.uploaded_at else "",
            "Tax Year": d.tax_year or "",
            "Review Status": (d.review_status or "").capitalize(),
            "Review Note": d.review_note or "",
        }
        if d.extracted_data:
            try:
                data = _json.loads(d.extracted_data) if isinstance(d.extracted_data, str) else d.extracted_data
                if isinstance(data, dict):
                    for k, v in data.items():
                        col = f"AI: {k}"
                        extracted_keys.add(col)
                        row[col] = str(v)
            except Exception:
                pass
        all_docs.append(row)

    base_cols = ["ID", "Table", "Doc Type", "Filename", "Uploaded At", "Tax Year", "Review Status", "Review Note"]
    all_cols = base_cols + sorted(list(extracted_keys))

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    # Header style
    header_fill  = PatternFill(fill_type="solid", fgColor="2C7A5B")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Alternate row fill
    alt_fill = PatternFill(fill_type="solid", fgColor="EAF4EF")

    # Write data
    for row_idx, doc in enumerate(all_docs, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, col_name in enumerate(all_cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=doc.get(col_name, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = thin_border
            if fill:
                cell.fill = fill

    # Auto-fit column widths (capped at 40)
    for col_idx, col_name in enumerate(all_cols, 1):
        max_len = max(len(col_name), *(len(str(doc.get(col_name, ""))) for doc in all_docs)) if all_docs else len(col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="user_{user_id}_documents.xlsx"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    return StreamingResponse(buf, headers=headers)

